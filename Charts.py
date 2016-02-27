from  openpyxl import (Workbook,load_workbook)


def extract_Ipeaks(device,c):
    wb = load_workbook('/home/aashish/DI/data/'+device+'.xlsx')                 #create instance of load_workbook
    work_sheet=wb['Sheet1']                                                      #select the required worksheet
    c=randrange(65,85) if c==0 else c+64                                        #if no trial specified choose any random trail
    data=[]                                                                     #list to store Ipeak values
    print '*'*39,chr(c),'*'*40
    for i in range(2,202):                                                      #copy values from cells and store it in data list
        data.append(work_sheet[chr(c)+str(i)].value)
    return data

def select_Ipeaks(data):
    i=0
    while(data[i+1]-data[i]<4 and i<(len(data)-2)):
        i+=1
    i-=2
    return data[i:]

devices=['Bulb_Cooler','Bulb_Fan']
#creating a new excel sheet
wb=Workbook()
ws=wb.active

for device in devices:
    ws['A1']='Number'
    raw_input("press any key")
    print device
    for x in range(2,62):
        ws['A'+str(x)]=str(x-1)
    for i in range(1,21):
        data=select_Ipeaks(extract_Ipeaks(device,i))
        print'+'*80
        print i
        print data
        print '_'*80
        ws[chr(i+65)+'1']='trial'+str(i)
        k=2
        temp=[]
        for j in range(0,len(data)):
            print i,' ',j,' ',k
            ws[chr(i+65)+str(k)]=data[j]-data[0]
            k+=1
            j+=1
            if(k>31):
                temp=data[j:]
                break
        # raw_input("Enter anything")
        temp=select_Ipeaks(temp)
        # print temp
        # raw_input("Enter anything")
        for j in range(0,len(temp)):
            print i,' ',j,' ',k
            ws[chr(i+65)+str(k)]=temp[j]-data[0]
            k+=1
            j+=1
            if(k>61):
                break
        wb.save('/home/aashish/DI/charts/'+device+'_50.xlsx')
print "WorkBooks created"
